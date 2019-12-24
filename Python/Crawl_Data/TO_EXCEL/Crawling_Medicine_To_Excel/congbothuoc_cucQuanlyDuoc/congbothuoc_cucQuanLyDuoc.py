import requests
# import the beautiful soup library
from bs4 import BeautifulSoup

import re

import xlwt
from xlrd import open_workbook
from xlutils.copy import copy as xl_copy
from openpyxl.styles import Font

import time
import sys

class ThuocBietDuoc:
  def __init__(self, url, post_url = "", name = ""):
    self.url = url
    self.post_url = post_url
    self.name = name

  def get_medicine_drug_groups(self, url):
    response = requests.get(self, timeout=10, stream=True)

    soup = BeautifulSoup(response.text, "html.parser")
    default = soup.find_all("a", class_="textpldl")
    # TODO: url nhóm thuốc
    urls = []

    # tên nhóm thuốc
    names = []
    for item in default:
      get_hrefs = item.get("href")
      name = item.get("title")
      urls.append("{}{}".format(url, get_hrefs[2:]))
      names.append(name)
    return urls, names

  def save_to_excel(self, sheet_name, data, row, column):
    rb = open_workbook(sheet_name, formatting_info=True)

    if rb is not None:
      r_sheet = rb.sheet_by_index(0)
      wb = copy(rb)
      sheet = wb.get_sheet(0)

      i = 0
      while i < len(data):
        writing = sheet.write(row, column, data[i])
        # print(len(sheet._Worksheet__rows)) # count row in sheet
        # print(r_sheet.nrows)
        wb.save(sheet_name)
        i = i + 1
        row = row + 1

  def save_home_to_excel(self, sheet_name, data, row, column):
    rb = open_workbook(sheet_name, formatting_info=True)

    if rb is not None:
      r_sheet = rb.sheet_by_index(0)
      wb = copy(rb)
      sheet = wb.get_sheet(0)

      writing = sheet.write(row, column, data)

      wb.save(sheet_name)

  def write_row_to_excel(sheet, text_data, row, column):
    sheet.write(row, column, text_data)

  def check_presence_find_next_tag_by_find(tag, tag_name):
    if tag is None:
      tag = ""
    else:
      tag = tag.find_next("{}".format(tag_name)).get_text(strip=True)

  def check_presence_find_next_div_by_find_all(tag_contents, text_info):
    for content in tag_contents:
      if content is None:
        content = ""
      else:
        text_info = (text_info.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()

  # def save_detail_drug_to_excel(data, detail_url):
  def save_detail_drug_to_excel(sheet, line, detail_url):
    default_url = "https://www.thuocbietduoc.com.vn"

    print("============== detail_url {}\t{}".format(line, detail_url))
    # while i < len(data):
    with requests.Session() as s:
      res = s.get(detail_url, timeout=20, stream=True)
      soup = BeautifulSoup(res.text,"lxml")

      find_soup = soup.find("article")
      # print("================= data", detail_url)
      ThuocBietDuoc.write_row_to_excel(sheet, detail_url, line, 0)

      # nhom_thuoc
      try:
        nhom_thuoc = find_soup.find('a', class_="textdetaillink")

        if not nhom_thuoc:
          nhom_thuoc = ""
        else:
          nhom_thuoc = nhom_thuoc.get("title")
        ThuocBietDuoc.write_row_to_excel(sheet, nhom_thuoc, line, 1)

        # Tên thuốc
        ten_thuoc = find_soup.h1
        if not ten_thuoc:
          ten_thuoc = ""
        else:
          ten_thuoc = ten_thuoc.get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, ten_thuoc, line, 2)

        # img
        img = find_soup.find("img", class_="imgdrg_lst")
        if not img:
          img = ""
        else:
          img = img.get("src")
        ThuocBietDuoc.write_row_to_excel(sheet, img, line, 3)

        # Dạng bào chế
        dang_bao_che = find_soup.find("span", class_="textdetailhead1", string=re.compile("Dạng bào chế"))
        if not dang_bao_che:
          dang_bao_che = ""
        else:
          dang_bao_che = dang_bao_che.find_next('span').get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, dang_bao_che, line, 4)

        #  Đóng gói
        dong_goi = find_soup.find("span", string=re.compile("Đóng gói"))
        if not dong_goi:
          dong_goi = ""
        else:
          dong_goi = dong_goi.find_next('span').get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, dong_goi, line, 5)


        # Thành phần
        thanh_phan = find_soup.find(string=re.compile("Thành phần"))
        if not thanh_phan:
          thanh_phan = ""
        else:
          thanh_phan = thanh_phan.find_next('span').get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, thanh_phan, line, 6)

        # SDK
        sdk = find_soup.find(string=re.compile("SĐK"))
        if not sdk:
          sdk = ""
        else:
          sdk = sdk.find_next('span').get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, sdk, line, 7)

        # Nhà sản xuất
        nha_san_xuat = find_soup.find(string=re.compile("Nhà sản xuất"))
        if not nha_san_xuat:
          nha_san_xuat = ""
        else:
          nha_san_xuat = nha_san_xuat.find_next('span').get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, nha_san_xuat, line, 8)

        # Nhà đăng ký
        nha_dang_ky = find_soup.find(string=re.compile("Nhà đăng ký"))
        if not nha_dang_ky:
          nha_dang_ky = ""
        else:
          nha_dang_ky = nha_dang_ky.find_next('span').get_text(strip=True)
        # print(">>>>>>>>>", nha_dang_ky)
        ThuocBietDuoc.write_row_to_excel(sheet, nha_dang_ky, line, 9)

        # Nhà phân phối
        nha_phan_phoi = find_soup.find(id="lblnhapp", string=re.compile("Nhà phân phối"))
        if not nha_phan_phoi:
          nha_phan_phoi = ""
        else:
          nha_phan_phoi = nha_phan_phoi.find_next('td').get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, nha_phan_phoi, line, 10)

        # chỉ định
        chi_dinhs = find_soup.find_all(string=re.compile("Chỉ định"))
        chi_dinh = ""
        # ThuocBietDuoc.check_presence_find_next_div_by_find_all(chi_dinhs, chi_dinh)
        for content in chi_dinhs:
          if content is None:
            content = ""
          else:
            chi_dinh = (chi_dinh.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()
        ThuocBietDuoc.write_row_to_excel(sheet, chi_dinh, line, 11)

        # Liều lượng - cách dùng
        lieuluong_cachdungs = find_soup.find_all(string=re.compile("Liều lượng - cách dùng"[1:].lower()))

        lieuluong_cachdung = ""
        for content in lieuluong_cachdungs:
          if content is None:
            content = ""
          else:
            lieuluong_cachdung = (lieuluong_cachdung.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()
        # ThuocBietDuoc.check_presence_find_next_div_by_find_all(lieuluong_cachdungs, lieuluong_cachdung)
        # print(">>>>>>>>>", lieuluong_cachdung)
        ThuocBietDuoc.write_row_to_excel(sheet, lieuluong_cachdung, line, 12)

        # chống chỉ định
        chong_chi_dinhs = find_soup.find_all(string=re.compile("Chống chỉ định"[1:].lower()))

        chong_chi_dinh = ""
        for content in chong_chi_dinhs:
          if content is None:
            content = ""
          else:
            chong_chi_dinh = (chong_chi_dinh.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()
        # ThuocBietDuoc.check_presence_find_next_div_by_find_all(chong_chi_dinhs, chong_chi_dinh)
        # print(">>>>>>>>>", chong_chi_dinh)
        ThuocBietDuoc.write_row_to_excel(sheet, chong_chi_dinh, line, 13)


        # Chú ý đề phòng
        chu_y_de_phong = find_soup.find("h2", class_="textdetailhead1", string=re.compile("Chú ý đề phòng"))
        if not chu_y_de_phong:
          chu_y_de_phong = ""
        else:
          chu_y_de_phong = chu_y_de_phong.find_next("div").get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, chu_y_de_phong, line, 14)

        ######### Thông tin thành phần
        # Dược lực
        duoc_luc = find_soup.find("h3", string=re.compile("Dược lực"))
        if not duoc_luc:
          duoc_luc = ""
        else:
          duoc_luc = duoc_luc.find_next("div").get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, duoc_luc, line, 15)

        # Dược động học
        duoc_dong_hoc = find_soup.find("h3", string=re.compile("Dược động học"))
        if not duoc_dong_hoc:
          duoc_dong_hoc = ""
        else:
          duoc_dong_hoc = duoc_dong_hoc.find_next("div").get_text(strip=True)
        ThuocBietDuoc.write_row_to_excel(sheet, duoc_dong_hoc, line, 16)

        # Tác dụng
        tac_dung = find_soup.find("h3", string="Tác dụng :")
        if not tac_dung:
          tac_dung = ""
        else:
          tac_dung = tac_dung.find_next("div").get_text(strip=True)
        # ThuocBietDuoc.check_presence_find_next_tag_by_find(tac_dung, "div")
        ThuocBietDuoc.write_row_to_excel(sheet, tac_dung, line, 17)

        # Tác dụng phụ
        tac_dung_phus = find_soup.find_all("section", id=re.compile("tac-dung-phu"), string=re.compile("Tác dụng phụ"[1:].lower()))
        tac_dung_phu = ""
        for content in tac_dung_phus:
          if content is None:
            content = ""
          else:
            tac_dung_phu = (tac_dung_phu.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()
        ThuocBietDuoc.write_row_to_excel(sheet, tac_dung_phu, line, 18)
        line = line + 1
      except:
        pass

  def create_headers_to_excel(sheet_name, excel_name_file):
    wb = xlwt.Workbook(encoding='utf-8', style_compression = 0)

    ws = wb.add_sheet(sheet_name)

    headers = ("url", "nhom_thuoc", "ten_thuoc", "img", "dang_bao_che", "dong_goi", "thanh_phan", "sdk", "nha_san_xuat", "nha_dang_ky", "nha_phan_phoi", "chi_dinh", "lieuluong_cachdung", "chong_chi_dinh", "chu_y_de_phong", "duoc_luc", "duoc_dong_hoc", "tac_dung", "tac_dung_phu")

    style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')
    for col, header_name in enumerate(headers):
      ws.row(0).write(col, header_name, style)

##### Get urls href base on navbar index page

default_url = "https://dichvucong.dav.gov.vn/congbothuoc/index"
# url = "https://www.thuocbietduoc.com.vn/thuoc/drgsearch.aspx"

print(">>>>>>>>>>>>>>>>>> Here We Go ")
from datetime import datetime

dt_now = datetime.now()
print("Start at: {}".format(dt_now))

# thongTinDangKyThuoc ( soQuyetDinh, dotCap)
# thongTinThuocCoBan (hoatchatChinh, hamLuong, dangBaoChe, dongGoi, tieuChuan, tuoiTho)
# congtySanXuat ("tenCongTySanXuat", "nuocSanXuat" , diaChiSanXuat)
# congTyDangKy (tenCongTyDangKy, diaChiDangKy, nuocDangKy)


# ####################################### Create file Excel ##################################
# wb = xlwt.Workbook(encoding="utf-8", style_compression=0)

# ws = wb.add_sheet("congbothuoc")

# headers = ("soDangKy", "tenThuoc", "hoatChatChinh", "hamLuong", "soQuyetDinh", "dotCap", "dangBaoChe", "dongGoi", "tieuChuan", "tuoiTho", "tenCongTyDangKy", "nuocDangKy", "diaChiDangKy", "tenCongTySanXuat", "nuocSanXuat", "diaChiSanXuat")

# style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')

# for col, header_name in enumerate(headers):
#   print("Col", col)
#   print("drug_info", header_name)
#   print(type(headers))
#   ws.row(0).write(col, header_name, style)
# i = n = 0
# while n < 39135:
# headers_req = requests.utils.default_headers()
# payload = {"SoDangKyThuoc": {}, "KichHoat": True, "skipCount": 0, "maxResultCount": 100, "sorting": None}
# res_post = requests.post("https://dichvucong.dav.gov.vn/api/services/app/soDangKy/GetAllPublicServerPaging", json=payload,headers=headers_req, timeout=10)

# items_thuoc = res_post.json()["result"]["items"]#[n]

# soDangKy = items_thuoc["soDangKy"]
# tenThuoc = items_thuoc["tenThuoc"]
# hoatChatChinh = items_thuoc["thongTinThuocCoBan"]["hoatChatChinh"]
# hamLuong = items_thuoc["thongTinThuocCoBan"]["hamLuong"]
# soQuyetDinh  = items_thuoc["thongTinDangKyThuoc"]["soQuyetDinh"]
# soQuyetDinh  = items_thuoc["thongTinDangKyThuoc"]["dotCap"]
# dangBaoChe = items_thuoc["thongTinThuocCoBan"]["dangBaoChe"]
# dongGoi = items_thuoc["thongTinThuocCoBan"]["dongGoi"]
# tieuChuan = items_thuoc["thongTinThuocCoBan"]["tieuChuan"]
# tuoiTho = items_thuoc["thongTinThuocCoBan"]["tuoiTho"]
# tenCongTyDangKy = items_thuoc["congTyDangKy"]["tenCongTyDangKy"]
# nuocDangKy = items_thuoc["congTyDangKy"]["nuocDangKy"]
# diaChiDangKy = items_thuoc["congTyDangKy"]["diaChiDangKy"]
# tenCongTySanXuat = items_thuoc["congTySanXuat"]["tenCongTySanXuat"]
# nuocSanXuat = items_thuoc["congTySanXuat"]["nuocSanXuat"]
# diaChiSanXuat = items_thuoc["congTySanXuat"]["diaChiSanXuat"]

# print(">>>>>>>>>>>>. 1\t{} 2\t{} 3\t{} 4\t{} 5\t{} 6\t{} 7\t{} 8\t{} 9\t{} 10\t{} 11\t{} 12\t{} 13\t{} 14\t{} 15\t{}".format(soDangKy, tenThuoc, hoatChatChinh, hamLuong, soQuyetDinh, dangBaoChe, dongGoi, tieuChuan, tuoiTho, tenCongTyDangKy, nuocDangKy, diaChiDangKy, tenCongTySanXuat, nuocSanXuat, diaChiSanXuat))
# print("line ", n)
  # n = n + 1
  # i = i + 1000




####################################### Create file Excel ##################################
book = xlwt.Workbook(encoding="utf-8", style_compression=0)

ws = book.add_sheet("congbothuoc", cell_overwrite_ok=True)

headers = ("soDangKy", "tenThuoc", "hoatChatChinh", "hamLuong", "soQuyetDinh", "dotCap", "dangBaoChe", "dongGoi", "tieuChuan", "tuoiTho", "tenCongTyDangKy", "nuocDangKy", "diaChiDangKy", "tenCongTySanXuat", "nuocSanXuat", "diaChiSanXuat")

style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')

for col, header_name in enumerate(headers):
  ws.row(0).write(col, header_name, style)

book.save("cucQuanLyDuoc.xls")

try:
  n = 0
  maxResultCount = 1000
  line = 1

  open_wb = open_workbook("cucQuanLyDuoc.xls", formatting_info=True)
  wb_copy = xl_copy(open_wb)

  sheet = wb_copy.add_sheet("cucQuanLyDuoc")
  headers = ("soDangKy", "tenThuoc", "hoatChatChinh", "hamLuong", "soQuyetDinh", "dotCap", "dangBaoChe", "dongGoi", "tieuChuan", "tuoiTho", "tenCongTyDangKy", "nuocDangKy", "diaChiDangKy", "tenCongTySanXuat", "nuocSanXuat", "diaChiSanXuat")

  style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')

  for col, header_name in enumerate(headers):
    sheet.row(0).write(col, header_name, style)

  headers_req = requests.utils.default_headers()
  payload = {"SoDangKyThuoc": {}, "KichHoat": True, "skipCount": 0, "maxResultCount": 100, "sorting": None}
  res = requests.post("https://dichvucong.dav.gov.vn/api/services/app/soDangKy/GetAllPublicServerPaging", json=payload,headers=headers_req, timeout=10)

  while n < res.json()["result"]["totalCount"]:
    headers_req = requests.utils.default_headers()
    payload = {"SoDangKyThuoc": {}, "KichHoat": True, "skipCount": n, "maxResultCount": maxResultCount, "sorting": None}
    res_post = requests.post("https://dichvucong.dav.gov.vn/api/services/app/soDangKy/GetAllPublicServerPaging", json=payload,headers=headers_req, timeout=10)
    items_thuoc = res_post.json()["result"]["items"][0]
    soDangKy = items_thuoc["soDangKy"]
    tenThuoc = items_thuoc["tenThuoc"]
    hoatChatChinh = items_thuoc["thongTinThuocCoBan"]["hoatChatChinh"]
    hamLuong = items_thuoc["thongTinThuocCoBan"]["hamLuong"]
    soQuyetDinh  = items_thuoc["thongTinDangKyThuoc"]["soQuyetDinh"]
    dotCap  = items_thuoc["thongTinDangKyThuoc"]["dotCap"]
    dangBaoChe = items_thuoc["thongTinThuocCoBan"]["dangBaoChe"]
    dongGoi = items_thuoc["thongTinThuocCoBan"]["dongGoi"]
    tieuChuan = items_thuoc["thongTinThuocCoBan"]["tieuChuan"]
    tuoiTho = items_thuoc["thongTinThuocCoBan"]["tuoiTho"]
    tenCongTyDangKy = items_thuoc["congTyDangKy"]["tenCongTyDangKy"]
    nuocDangKy = items_thuoc["congTyDangKy"]["nuocDangKy"]
    diaChiDangKy = items_thuoc["congTyDangKy"]["diaChiDangKy"]
    tenCongTySanXuat = items_thuoc["congTySanXuat"]["tenCongTySanXuat"]
    nuocSanXuat = items_thuoc["congTySanXuat"]["nuocSanXuat"]
    diaChiSanXuat = items_thuoc["congTySanXuat"]["diaChiSanXuat"]

    print(">>>>>>>>>>>>. 1 \t{} 2 \t{} 3 \t{} 4 \t{} 5 \t{} 6 \t{} 7 \t{} 8 \t{} 9 \t{} 10 \t{} 11 \t{} 12\t{} 13\t{} 14\t{} 15\t{}".format(soDangKy, tenThuoc, hoatChatChinh, hamLuong, soQuyetDinh, dangBaoChe, dongGoi, tieuChuan, tuoiTho, tenCongTyDangKy, nuocDangKy, diaChiDangKy, tenCongTySanXuat, nuocSanXuat, diaChiSanXuat))

    drug_infos = (soDangKy, tenThuoc, hoatChatChinh, hamLuong, soQuyetDinh, dotCap, dangBaoChe, dongGoi, tieuChuan, tuoiTho, tenCongTyDangKy, nuocDangKy, diaChiDangKy, tenCongTySanXuat, nuocSanXuat, diaChiSanXuat)

    for col, drug_info in enumerate(drug_infos):
      sheet.row(line).write(col, drug_info)

    print("line: ", n)

    n = n + 1
    line = line + 1
  wb_copy.save("cucQuanLyDuoc.xls")
except ValueError as e:
  print("error here", e)

end = (datetime.now() - dt_now)

print("We spent {} to done this".format(end))
