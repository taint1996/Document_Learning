import requests
# import the beautiful soup library
from bs4 import BeautifulSoup

import re

import xlwt
from xlrd import open_workbook
from xlutils.copy import copy as xl_copy
from openpyxl.styles import Font

import time
import sys

class ThuocBietDuoc:
  def __init__(self, url, post_url = "", name = ""):
    self.url = url
    self.post_url = post_url
    self.name = name

  def get_medicine_drug_groups(self, url):
    response = requests.get(self, timeout=10, stream=True)

    soup = BeautifulSoup(response.text, "html.parser")
    default = soup.find_all("a", class_="textpldl")
    # TODO: url nhóm thuốc
    urls = []

    # tên nhóm thuốc
    names = []
    for item in default:
      get_hrefs = item.get("href")
      name = item.get("title")
      urls.append("{}{}".format(url, get_hrefs[2:]))
      names.append(name)
    return urls, names

  def save_to_excel(self, sheet_name, data, row, column):
    rb = open_workbook(sheet_name, formatting_info=True)

    if rb is not None:
      r_sheet = rb.sheet_by_index(0)
      wb = copy(rb)
      sheet = wb.get_sheet(0)

      i = 0
      while i < len(data):
        writing = sheet.write(row, column, data[i])
        # print(len(sheet._Worksheet__rows)) # count row in sheet
        # print(r_sheet.nrows)
        wb.save(sheet_name)
        i = i + 1
        row = row + 1

  def save_home_to_excel(self, sheet_name, data, row, column):
    rb = open_workbook(sheet_name, formatting_info=True)

    if rb is not None:
      r_sheet = rb.sheet_by_index(0)
      wb = copy(rb)
      sheet = wb.get_sheet(0)

      writing = sheet.write(row, column, data)

      wb.save(sheet_name)

  def write_row_to_excel(sheet, text_data, row, column):
    sheet.write(row, column, text_data)

  def check_presence_find_next_tag_by_find(tag, tag_name):
    if tag is None:
      tag = ""
    else:
      tag = tag.find_next("{}".format(tag_name)).get_text(strip=True)

  def check_presence_find_next_div_by_find_all(tag_contents, text_info):
    for content in tag_contents:
      if content is None:
        content = ""
      else:
        text_info = (text_info.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()

  # def save_detail_drug_to_excel(data, detail_url):
  def save_detail_drug_to_excel(sheet, line, detail_url):
    default_url = "https://www.thuocbietduoc.com.vn"

    print("============== detail_url {}\t{}".format(line, detail_url))
    # while i < len(data):
    with requests.Session() as s:
      res = s.get(detail_url, timeout=5, stream=True)
      soup = BeautifulSoup(res.text,"lxml")

      find_soup = soup.find("article")
      # print("================= data", detail_url)
      ThuocBietDuoc.write_row_to_excel(sheet, detail_url, line, 0)

      # nhom_thuoc
      nhom_thuoc = find_soup.find('a', class_="textdetaillink")
      if not nhom_thuoc:
        nhom_thuoc = ""
      else:
        nhom_thuoc = nhom_thuoc.get("title")
      ThuocBietDuoc.write_row_to_excel(sheet, nhom_thuoc, line, 1)

      # Tên thuốc
      ten_thuoc = find_soup.h1
      if not ten_thuoc:
        ten_thuoc = ""
      else:
        ten_thuoc = ten_thuoc.get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, ten_thuoc, line, 2)

      # img
      img = find_soup.find("img", class_="imgdrg_lst")
      if not img:
        img = ""
      else:
        img = img.get("src")
      ThuocBietDuoc.write_row_to_excel(sheet, img, line, 3)

      # Dạng bào chế
      dang_bao_che = find_soup.find("span", class_="textdetailhead1", string=re.compile("Dạng bào chế"))
      if not dang_bao_che:
        dang_bao_che = ""
      else:
        dang_bao_che = dang_bao_che.find_next('span').get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, dang_bao_che, line, 4)

      #  Đóng gói
      dong_goi = find_soup.find("span", string=re.compile("Đóng gói"))
      if not dong_goi:
        dong_goi = ""
      else:
        dong_goi = dong_goi.find_next('span').get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, dong_goi, line, 5)


      # Thành phần
      thanh_phan = find_soup.find(string=re.compile("Thành phần"))
      if not thanh_phan:
        thanh_phan = ""
      else:
        thanh_phan = thanh_phan.find_next('span').get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, thanh_phan, line, 6)

      # SDK
      sdk = find_soup.find(string=re.compile("SĐK"))
      if not sdk:
        sdk = ""
      else:
        sdk = sdk.find_next('span').get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, sdk, line, 7)

      # Nhà sản xuất
      nha_san_xuat = find_soup.find(string=re.compile("Nhà sản xuất"))
      if not nha_san_xuat:
        nha_san_xuat = ""
      else:
        nha_san_xuat = nha_san_xuat.find_next('span').get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, nha_san_xuat, line, 8)

      # Nhà đăng ký
      nha_dang_ky = find_soup.find(string=re.compile("Nhà đăng ký"))
      if not nha_dang_ky:
        nha_dang_ky = ""
      else:
        nha_dang_ky = nha_dang_ky.find_next('span').get_text(strip=True)
      # print(">>>>>>>>>", nha_dang_ky)
      ThuocBietDuoc.write_row_to_excel(sheet, nha_dang_ky, line, 9)

      # Nhà phân phối
      nha_phan_phoi = find_soup.find(id="lblnhapp", string=re.compile("Nhà phân phối"))
      if not nha_phan_phoi:
        nha_phan_phoi = ""
      else:
        nha_phan_phoi = nha_phan_phoi.find_next('td').get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, nha_phan_phoi, line, 10)

      # chỉ định
      chi_dinhs = find_soup.find_all(string=re.compile("Chỉ định"))
      chi_dinh = ""
      # ThuocBietDuoc.check_presence_find_next_div_by_find_all(chi_dinhs, chi_dinh)
      for content in chi_dinhs:
        if content is None:
          content = ""
        else:
          chi_dinh = (chi_dinh.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()
      ThuocBietDuoc.write_row_to_excel(sheet, chi_dinh, line, 11)

      # Liều lượng - cách dùng
      lieuluong_cachdungs = find_soup.find_all(string=re.compile("Liều lượng - cách dùng"[1:].lower()))

      lieuluong_cachdung = ""
      for content in lieuluong_cachdungs:
        if content is None:
          content = ""
        else:
          lieuluong_cachdung = (lieuluong_cachdung.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()
      # ThuocBietDuoc.check_presence_find_next_div_by_find_all(lieuluong_cachdungs, lieuluong_cachdung)
      # print(">>>>>>>>>", lieuluong_cachdung)
      ThuocBietDuoc.write_row_to_excel(sheet, lieuluong_cachdung, line, 12)

      # chống chỉ định
      chong_chi_dinhs = find_soup.find_all(string=re.compile("Chống chỉ định"[1:].lower()))

      chong_chi_dinh = ""
      for content in chong_chi_dinhs:
        if content is None:
          content = ""
        else:
          chong_chi_dinh = (chong_chi_dinh.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()
      # ThuocBietDuoc.check_presence_find_next_div_by_find_all(chong_chi_dinhs, chong_chi_dinh)
      # print(">>>>>>>>>", chong_chi_dinh)
      ThuocBietDuoc.write_row_to_excel(sheet, chong_chi_dinh, line, 13)


      # Chú ý đề phòng
      chu_y_de_phong = find_soup.find("h2", class_="textdetailhead1", string=re.compile("Chú ý đề phòng"))
      if not chu_y_de_phong:
        chu_y_de_phong = ""
      else:
        chu_y_de_phong = chu_y_de_phong.find_next("div").get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, chu_y_de_phong, line, 14)

      ######### Thông tin thành phần
      # Dược lực
      duoc_luc = find_soup.find("h3", string=re.compile("Dược lực"))
      if not duoc_luc:
        duoc_luc = ""
      else:
        duoc_luc = duoc_luc.find_next("div").get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, duoc_luc, line, 15)

      # Dược động học
      duoc_dong_hoc = find_soup.find("h3", string=re.compile("Dược động học"))
      if not duoc_dong_hoc:
        duoc_dong_hoc = ""
      else:
        duoc_dong_hoc = duoc_dong_hoc.find_next("div").get_text(strip=True)
      ThuocBietDuoc.write_row_to_excel(sheet, duoc_dong_hoc, line, 16)

      # Tác dụng
      tac_dung = find_soup.find("h3", string="Tác dụng :")
      if not tac_dung:
        tac_dung = ""
      else:
        tac_dung = tac_dung.find_next("div").get_text(strip=True)
      # ThuocBietDuoc.check_presence_find_next_tag_by_find(tac_dung, "div")
      ThuocBietDuoc.write_row_to_excel(sheet, tac_dung, line, 17)

      # Tác dụng phụ
      tac_dung_phus = find_soup.find_all("section", id=re.compile("tac-dung-phu"), string=re.compile("Tác dụng phụ"[1:].lower()))
      tac_dung_phu = ""
      for content in tac_dung_phus:
        if content is None:
          content = ""
        else:
          tac_dung_phu = (tac_dung_phu.strip() + "\n" + content.find_next("div").get_text(strip=True)).strip()
      ThuocBietDuoc.write_row_to_excel(sheet, tac_dung_phu, line, 18)
      line = line + 1

  def create_headers_to_excel(sheet_name, excel_name_file):
    wb = xlwt.Workbook(encoding='utf-8', style_compression = 0)

    ws = wb.add_sheet(sheet_name)

    headers = ("url", "nhom_thuoc", "ten_thuoc", "img", "dang_bao_che", "dong_goi", "thanh_phan", "sdk", "nha_san_xuat", "nha_dang_ky", "nha_phan_phoi", "chi_dinh", "lieuluong_cachdung", "chong_chi_dinh", "chu_y_de_phong", "duoc_luc", "duoc_dong_hoc", "tac_dung", "tac_dung_phu")

    style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')
    for col, header_name in enumerate(headers):
      ws.row(0).write(col, header_name, style)

##### Get urls href base on navbar index page

default_url = "https://www.thuocbietduoc.com.vn"
url = "https://www.thuocbietduoc.com.vn/thuoc/drgsearch.aspx"

print(">>>>>>>>>>>>>>>>>> Here We Go ")
from datetime import datetime

dt_now = datetime.now()
print("Start at: {}".format(dt_now))

# ######################################## List URL Nhóm thuốc ########################################
medicine_group_urls = ThuocBietDuoc.get_medicine_drug_groups(url, default_url)





# i = 0
# total = []
# urls_medicine_groups = []

# sheet_drug_group_name = medicine_group_urls[0][i].split("/")[-2] # nhom-thuoc-1-0

# excel_name = default_url.split(".")[1] # thuocbietduoc
# excel_name_file = "{}.xls".format(excel_name) # thuocbietduoc.xls

# while i < (len(medicine_group_urls[0]) - 1):
#   book = xlwt.Workbook(encoding='utf-8', style_compression = 0)
#   sheet = book.add_sheet(sheet_drug_group_name, cell_overwrite_ok = True)

#   headers = ("url", "nhom_thuoc", "ten_thuoc", "img", "dang_bao_che", "dong_goi", "thanh_phan", "sdk", "nha_san_xuat", "nha_dang_ky", "nha_phan_phoi", "chi_dinh", "lieuluong_cachdung", "chong_chi_dinh", "chu_y_de_phong", "duoc_luc", "duoc_dong_hoc", "tac_dung", "tac_dung_phu")

#   style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')
#   for col, header in enumerate(headers):
#     sheet.row(0).write(col, header, style)

#   line = 1
#   with requests.Session() as s:
#     s.headers={
#                 "User-Agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36"
#               }
#     n = 1
#     res = s.get(medicine_group_urls[0][i], headers=s.headers, timeout=20, stream=True)
#     soup = BeautifulSoup(res.text,"lxml")

#     while True:
#       payload = {item['name']:item.get('value','') for item in soup.select("input[name]")}
#       payload['__VIEWSTATEGENERATOR'] = '336E2262'
#       payload['Drugnews1$Gr1'] = 'Rdrgnm'
#       payload['page'] = n
#       payload['currentView'] = '1'
#       req = s.post(medicine_group_urls[0][i], data=payload,headers=s.headers)
#       soup_obj = BeautifulSoup(req.text,"html.parser")

#       print("====== page n", n)

#       find_soup = soup_obj.find("table", attrs={"class": "text2", "id": "Tabl1"})

#       find_a_href = find_soup.find_all("a", class_="textlink01_v")

#       for url in find_a_href:
#         url_href = "{}{}".format(default_url, url.get("href")[2:])
#         name_drug = url_href[:-5].split("/")[-1] # Get name drug

#         # check unique url href #TODO: check later
#         ############### name drug use to uniq drug
#         if url_href not in urls_medicine_groups:# and name_drug not in urls_medicine_groups:
#           urls_medicine_groups.append(url_href)
#           ################################################### Save record to Excel #########################################
#           ThuocBietDuoc.save_detail_drug_to_excel(sheet, line, url_href)
#           line = line + 1
#       print("len medicine", len(urls_medicine_groups))

#       if len(find_a_href) == 0:
#         book.save(excel_name_file)

#         rb = open_workbook(excel_name_file, formatting_info=True)
#         wb = xl_copy(rb)
#         sheet_names = rb.sheet_names()

#         if IndexError:
#           break

#         if sheet_drug_group_name in sheet_names:
#           sheet_drug_group_name_1 = medicine_group_urls[0][i+1].split("/")[-2] # Tên nhóm thuốc

#           add_sheet = wb.add_sheet(sheet_drug_group_name_1)

#           headers = ("url", "nhom_thuoc", "ten_thuoc", "img", "dang_bao_che", "dong_goi", "thanh_phan", "sdk", "nha_san_xuat", "nha_dang_ky", "nha_phan_phoi", "chi_dinh", "lieuluong_cachdung", "chong_chi_dinh", "chu_y_de_phong", "duoc_luc", "duoc_dong_hoc", "tac_dung", "tac_dung_phu")

#           style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')
#           for col, header in enumerate(headers):
#             add_sheet.row(0).write(col, header, style)

#           # with requests.Session() as s:
#           #   s.headers={
#           #               "User-Agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36"
#           #             }
#             n = 1
#             res = s.get(medicine_group_urls[0][i], headers=s.headers, timeout=20, stream=True)
#             soup = BeautifulSoup(res.text,"lxml")

#             while True:
#               payload = {item['name']:item.get('value','') for item in soup.select("input[name]")}
#               payload['__VIEWSTATEGENERATOR'] = '336E2262'
#               payload['Drugnews1$Gr1'] = 'Rdrgnm'
#               payload['page'] = n
#               payload['currentView'] = '1'
#               req = s.post(medicine_group_urls[0][i], data=payload,headers=s.headers)
#               soup_obj = BeautifulSoup(req.text,"html.parser")

#               find_soup = soup_obj.find("table", attrs={"class": "text2", "id": "Tabl1"})

#               find_a_href = find_soup.find_all("a", class_="textlink01_v")
#               if len(find_a_href) == 0:
#                 break

#               for url in find_a_href:
#                 url_href = "{}{}".format(default_url, url.get("href")[2:])
#                 name_drug = url_href[:-5].split("/")[-1] # Get name drug

#                 # check unique url href #TODO: check later
#                 ############### name drug use to uniq drug
#                 if url_href not in urls_medicine_groups:# and name_drug not in urls_medicine_groups:
#                   urls_medicine_groups.append(url_href)
#                   ################################################### Save record to Excel #########################################
#                   ThuocBietDuoc.save_detail_drug_to_excel(add_sheet, line, url_href)
#                   line = line + 1
#               n = n + 1
#               print("Go next !")
#         wb.save(excel_name_file)
#         break
#       n = n + 1
#   i = i + 1

# rb = open_workbook(excel_name_file, formatting_info=True)

# wb = xl_copy(rb)
# sheet_names = rb.sheet_names()
# sheet_drug_group_name_1 = medicine_group_urls[0][-1].split("/")[-2] # Tên nhóm thuốc

# add_sheet = wb.add_sheet(sheet_drug_group_name_1)

# headers = ("url", "nhom_thuoc", "ten_thuoc", "img", "dang_bao_che", "dong_goi", "thanh_phan", "sdk", "nha_san_xuat", "nha_dang_ky", "nha_phan_phoi", "chi_dinh", "lieuluong_cachdung", "chong_chi_dinh", "chu_y_de_phong", "duoc_luc", "duoc_dong_hoc", "tac_dung", "tac_dung_phu")

# style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')
# for col, header in enumerate(headers):
#   add_sheet.row(0).write(col, header, style)

# line = 1
# with requests.Session() as s:
#   s.headers={
#               "User-Agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36"
#             }
#   n = 1
#   res = s.get(medicine_group_urls[0][-1], headers=s.headers, timeout=20, stream=True)
#   soup = BeautifulSoup(res.text,"lxml")

#   while True:
#     payload = {item['name']:item.get('value','') for item in soup.select("input[name]")}
#     payload['__VIEWSTATEGENERATOR'] = '336E2262'
#     payload['Drugnews1$Gr1'] = 'Rdrgnm'
#     payload['page'] = n
#     payload['currentView'] = '1'
#     req = s.post(medicine_group_urls[0][-1], data=payload,headers=s.headers)
#     soup_obj = BeautifulSoup(req.text,"html.parser")

#     find_soup = soup_obj.find("table", attrs={"class": "text2", "id": "Tabl1"})

#     find_a_href = find_soup.find_all("a", class_="textlink01_v")
#     if len(find_a_href) == 0:
#       break
#     if find_a_href is None:
#       continue
#     for url in find_a_href:
#       url_href = "{}{}".format(default_url, url.get("href")[2:])
#       name_drug = url_href[:-5].split("/")[-1] # Get name drug

#       # check unique url href #TODO: check later
#       ############### name drug use to uniq drug
#       if url_href not in urls_medicine_groups:# and name_drug not in urls_medicine_groups:
#         urls_medicine_groups.append(url_href)
#         ################################################### Save record to Excel #########################################
#         ThuocBietDuoc.save_detail_drug_to_excel(add_sheet, line, url_href)
#         line = line + 1
#     n = n + 1
# wb.save(excel_name_file)



























urls_medicine_groups = []

sheet_first_drug_group_name = medicine_group_urls[0][0].split("/")[-2] # nhom-thuoc-1-0

excel_name = default_url.split(".")[1] # thuocbietduoc
excel_name_file = "{}.xls".format(excel_name) # thuocbietduoc.xls

book = xlwt.Workbook(encoding='utf-8', style_compression = 0)
sheet = book.add_sheet(sheet_first_drug_group_name, cell_overwrite_ok = True)

headers = ("url", "nhom_thuoc", "ten_thuoc", "img", "dang_bao_che", "dong_goi", "thanh_phan", "sdk", "nha_san_xuat", "nha_dang_ky", "nha_phan_phoi", "chi_dinh", "lieuluong_cachdung", "chong_chi_dinh", "chu_y_de_phong", "duoc_luc", "duoc_dong_hoc", "tac_dung", "tac_dung_phu")

style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')
for col, header in enumerate(headers):
  sheet.row(0).write(col, header, style)

# with requests.Session() as s:
#   s.headers={
#               "User-Agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36"
#             }
headers = requests.utils.default_headers()
headers['User-Agent'] = 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36'

res = requests.get(medicine_group_urls[0][0], headers=headers, timeout=5, stream=True)
soup = BeautifulSoup(res.text,"lxml")
n = line = 1
while True:
  payload = {item['name']:item.get('value','') for item in soup.select("input[name]")}
  payload['__VIEWSTATEGENERATOR'] = '336E2262'
  payload['Drugnews1$Gr1'] = 'Rdrgnm'
  payload['page'] = n
  payload['currentView'] = '1'
  req = requests.post(medicine_group_urls[0][0], data=payload,headers=headers)
  soup_obj = BeautifulSoup(req.text,"html.parser")

  find_soup = soup_obj.find("table", attrs={"class": "text2", "id": "Tabl1"})

  find_a_href = find_soup.find_all("a", class_="textlink01_v")
  if len(find_a_href) == 0:
    break

  if find_a_href is None:
    continue

  for url in find_a_href:
    url_href = "{}{}".format(default_url, url.get("href")[2:])
    name_drug_aspx = url_href.split("/")[-1] # Get name drug
    # check unique url href #TODO: check later
    ############### name drug use to uniq drug
    if url_href not in urls_medicine_groups and name_drug_aspx not in url_href:
      urls_medicine_groups.append(url_href)
      ################################################### Save record to Excel #########################################
      ThuocBietDuoc.save_detail_drug_to_excel(sheet, line, url_href)
      line = line + 1
  n = n + 1

book.save(excel_name_file)

print("pass here 1")

i = 1
while i <= (len(medicine_group_urls[0])):
  print("hello there")
  rb = open_workbook(excel_name_file, formatting_info=True)
  wb = xl_copy(rb)
  sheet_names = rb.sheet_names()

  sheet_drug_group_name = medicine_group_urls[0][i].split("/")[-2] # Tên nhóm thuốc
  if sheet_drug_group_name not in sheet_names:
    line = 1

    print("whats'up")
    print(sheet_drug_group_name)
    print(sheet_names)
    add_sheet = wb.add_sheet(sheet_drug_group_name)

    headers = ("url", "nhom_thuoc", "ten_thuoc", "img", "dang_bao_che", "dong_goi", "thanh_phan", "sdk", "nha_san_xuat", "nha_dang_ky", "nha_phan_phoi", "chi_dinh", "lieuluong_cachdung", "chong_chi_dinh", "chu_y_de_phong", "duoc_luc", "duoc_dong_hoc", "tac_dung", "tac_dung_phu")

    style = xlwt.easyxf('align: vert centre, horiz centre; font: bold on')
    for col, header in enumerate(headers):
      add_sheet.row(0).write(col, header, style)

    n = 1

    # with requests.Session() as s:
    #   s.headers={
    #               "User-Agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36"
    #             }

    headers = requests.utils.default_headers()
    headers['User-Agent'] = 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36'
    res = requests.get(medicine_group_urls[0][i], headers=headers, timeout=5, stream=True)
    soup = BeautifulSoup(res.text,"lxml")

    while True:
      payload = {item['name']:item.get('value','') for item in soup.select("input[name]")}
      payload['__VIEWSTATEGENERATOR'] = '336E2262'
      payload['Drugnews1$Gr1'] = 'Rdrgnm'
      payload['page'] = n
      payload['currentView'] = '1'
      req = requests.post(medicine_group_urls[0][i], data=payload, headers=headers)
      soup_obj = BeautifulSoup(req.text,"html.parser")

      find_soup = soup_obj.find("table", attrs={"class": "text2", "id": "Tabl1"})

      find_a_href = find_soup.find_all("a", class_="textlink01_v")

      if len(find_a_href) == 0:
        break

      if find_a_href is None:
        continue

      for url in find_a_href:
        url_href = "{}{}".format(default_url, url.get("href")[2:])
        name_drug_aspx = url_href.split("/")[-1] # Get name drug
        # check unique url href #TODO: check later
        ############### name drug use to uniq drug
        if url_href not in urls_medicine_groups and name_drug_aspx not in url_href:
          urls_medicine_groups.append(url_href)
          ################################################### Save record to Excel #########################################
          ThuocBietDuoc.save_detail_drug_to_excel(add_sheet, line, url_href)
          line = line + 1
      n = n + 1
      print("Go next !")
  wb.save(excel_name_file)
  i = i + 1

end = (datetime.now() - dt_now)

print("We spent {} to done this".format(end))
